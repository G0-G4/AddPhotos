import PySimpleGUI as sg
import os
import shutil
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.drawing.image import Image
from PIL import Image as pilim
from PIL import ImageFile
import time
import traceback

ImageFile.LOAD_TRUNCATED_IMAGES = True

H = 0.75
W = 0.143
i = 0
DIR = 'thisisarandomname4r8r4u83rhuwrgniwrghirwhuwrg'

def add(image_name, cell, of, ws, size, quality):
    global H, W, i
    try:
        img = pilim.open(image_name)
        img = img.convert('RGB')
        img.save(f'{DIR}/img{i}.jpg', quality = quality)
        img = Image(f'{DIR}/img{i}.jpg')
        k = img.height / size
        img.height = size
        img.width /= k
        anchor = cell.offset(0, of)

        ws.row_dimensions[anchor.row].height = img.height * H
        ws.column_dimensions[get_column_letter(anchor.column)].width = img.width * W
        ws.add_image(img, anchor.coordinate)
        i += 1
    except Exception:
        traceback.print_exc()

def check_image(file):
    return (file.endswith('png')  or
           file.endswith('PNG')  or
           file.endswith('jpg')  or
           file.endswith('JPG')  or
           file.endswith('JPEG') or
           file.endswith('jpeg'))




def main(file_name, folder, size, quality, row, col, offset):
    wb = load_workbook(file_name)
    ws = wb.active

    # creating dict of cells with code as a key
    print('creating dict of codes')
    cells = {}
    for row in ws.iter_rows(min_row=row, min_col = col, max_col=col, max_row=ws.max_row):
        cells[str(row[0].value)] = row[0]

    # searching and inserting images
    print('searching and inserting images')
    for path, names, files in os.walk(folder):
        for code, cell in cells.items(): # also iterates over already filled codes
            if path.endswith(code):
                of = offset
                for file in files:
                    if check_image(file):
                        image_name = os.path.join(path, file)
                        print(image_name)
                        add(image_name, cell, of, ws, size, quality)
                        of += 1
                    
            of = offset
            for file in files:
                if file.startswith(code):
                    if check_image(file):
                        image_name = os.path.join(path, file)
                        print(image_name)
                        add(image_name, cell, of, ws, size, quality)
                        of += 1

    print('writing')
    wb.save(file_name)
    wb.close()
    print('DONE')


sliders = [
    [sg.Text('размер фотографий')],
    [sg.Slider(key = '-slider-', range = (100, 5000), default_value=1000, orientation = 'h', resolution=100)],
    [sg.Text('качество')],
    [sg.Slider(key = '-quality-', range = (5, 95), default_value=50, orientation = 'h', resolution=5)],
]

sett = [
    [sg.Input('A', size=10, key='-column-'), sg.T('столбец с кодами')],
    [sg.Input('E', size=10, key='-photo_column-'), sg.T('столбец для фото')],
    [sg.Input(2, size=10, key='-row-'), sg.T('ряд начала')]
]
settings = [
    [sg.Column(sett)]
]

layout = [
    [sg.Text('файл с кодами')],
    [sg.Input(key = '-file-'), sg.FileBrowse(target='-file-')],
    [sg.Text('папка с фотографиями')],
    [sg.Input(key = '-folder-'), sg.FolderBrowse()],
    [sg.Column(sliders), sg.Column(settings)],
    [sg.OK(), sg.Cancel()]]

window = sg.Window('', layout)
while True:
    event, values = window.read()
    if event == sg.WIN_CLOSED or event == 'Cancel':
        break
    if event == 'OK':
        try:
            start = time.time()
            if not os.path.exists(DIR):
                os.mkdir(DIR)
            file = values['-file-']
            folder = values['-folder-']
            size = int(values['-slider-'])
            q = int(values['-quality-'])
            col_c = values['-column-']
            try:
                col_c = column_index_from_string(col_c)
            except Exception as e:
                sg.popup(e)
                continue
            col_ph = values['-photo_column-']
            try:
                col_ph = column_index_from_string(col_ph)
            except Exception as e:
                sg.popup(e)
                continue
            row = values['-row-']
            if row.isdigit() and int(row) > 1:
                row = int(values['-row-'])
            else:
                sg.popup('ряд должен быть числом > 1')
                continue
            print(file, folder)
            if not os.path.exists(file) or not os.path.exists(folder):
                sg.popup('выберите файлы')
                continue
            print(col_ph - col_c, file)
            main(file, folder, size, q, row, col_c, col_ph - col_c)
            sg.popup('Done!')
            print(f"took {time.time() - start} seconds")
            shutil.rmtree(DIR)
        except Exception as e:
            sg.popup(e)
            traceback.print_exc()


window.close()